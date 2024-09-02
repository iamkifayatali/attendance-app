from openpyxl import load_workbook, Workbook

# Load the workbooks
human = load_workbook('/content/drive/MyDrive/attendance/human.xlsx') #firstfilelocation
machine = load_workbook('/content/drive/MyDrive/attendance/machine.xlsx')  #2ndfilelocation

# Get the active sheets
sheet1 = human.active
sheet2 = machine.active

# Create a new workbook
combined_wb = Workbook()
combined_sheet = combined_wb.active

# Copy data from sheet1
for row in sheet1.iter_rows(values_only=True):
  combined_sheet.append(row)

# Copy data from sheet2
for row in sheet2.iter_rows(values_only=True):
  combined_sheet.append(row)

# Save the combined workbook
combined_wb.save('/content/drive/MyDrive/attendance/combined.xlsx')